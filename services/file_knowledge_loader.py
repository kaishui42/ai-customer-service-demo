"""从 knowledge/ 目录动态加载本地知识文件（报价表等）。"""

from __future__ import annotations

import csv
import re
from dataclasses import dataclass
from pathlib import Path

import config

SUPPORTED_EXTENSIONS = {".txt", ".md", ".csv", ".xlsx"}
MAX_CONTEXT_CHARS = 18000
MAX_FILE_CHARS = 12000


@dataclass
class LoadedDocument:
    filename: str
    source_path: Path
    content: str
    mtime: float


@dataclass
class FileKnowledgeHit:
    filename: str
    excerpt: str
    score: int


class FileKnowledgeLoader:
    """扫描并解析 knowledge/ 目录下的知识文件。"""

    def __init__(self, knowledge_dir: Path | None = None) -> None:
        self.knowledge_dir = knowledge_dir or config.KNOWLEDGE_DIR
        self.knowledge_dir.mkdir(parents=True, exist_ok=True)
        self.documents: list[LoadedDocument] = []

    def reload(self) -> list[LoadedDocument]:
        """重新读取目录下所有支持格式的文件。"""
        self.documents = []

        for path in sorted(self.knowledge_dir.iterdir()):
            if not path.is_file():
                continue
            if path.name.startswith("."):
                continue
            if path.suffix.lower() not in SUPPORTED_EXTENSIONS:
                continue

            try:
                content = self._read_file(path)
            except Exception as exc:
                content = f"（文件读取失败：{exc}）"

            if not content.strip():
                continue

            if len(content) > MAX_FILE_CHARS:
                content = content[:MAX_FILE_CHARS] + "\n…（内容过长，已截断）"

            self.documents.append(
                LoadedDocument(
                    filename=path.name,
                    source_path=path,
                    content=content,
                    mtime=path.stat().st_mtime,
                )
            )

        return self.documents

    def has_documents(self) -> bool:
        return bool(self.documents)

    def filenames(self) -> list[str]:
        return [doc.filename for doc in self.documents]

    def search(self, query: str, top_k: int = 8) -> list[FileKnowledgeHit]:
        """按关键词在已加载文件中检索相关片段。"""
        if not query.strip() or not self.documents:
            return []

        terms = self._extract_terms(query)
        if not terms:
            return []

        hits: list[FileKnowledgeHit] = []

        for doc in self.documents:
            for line in doc.content.splitlines():
                stripped = line.strip()
                if not stripped or stripped.startswith("（"):
                    continue

                score = sum(1 for term in terms if term in stripped.lower())
                if score > 0:
                    hits.append(
                        FileKnowledgeHit(
                            filename=doc.filename,
                            excerpt=stripped,
                            score=score,
                        )
                    )

        hits.sort(key=lambda item: item.score, reverse=True)
        return hits[:top_k]

    def build_reference_context(
        self,
        query: str,
        hits: list[FileKnowledgeHit] | None = None,
    ) -> str:
        """组装供大模型参考的本地文件上下文。"""
        if not self.documents:
            return ""

        parts: list[str] = []
        matched_hits = hits if hits is not None else self.search(query)

        if matched_hits:
            parts.append("=== 检索到的相关片段 ===")
            for hit in matched_hits:
                parts.append(f"[{hit.filename}] {hit.excerpt}")

        parts.append("=== 本地文件完整参考 ===")
        for doc in self.documents:
            parts.append(f"--- 文件：{doc.filename} ---")
            parts.append(doc.content)

        context = "\n".join(parts)
        if len(context) > MAX_CONTEXT_CHARS:
            context = context[:MAX_CONTEXT_CHARS] + "\n…（参考内容过长，已截断）"
        return context

    def _read_file(self, path: Path) -> str:
        suffix = path.suffix.lower()

        if suffix in {".txt", ".md"}:
            return path.read_text(encoding="utf-8", errors="replace")

        if suffix == ".csv":
            return self._read_csv(path)

        if suffix == ".xlsx":
            return self._read_xlsx(path)

        return ""

    def _read_csv(self, path: Path) -> str:
        lines: list[str] = []
        with path.open(encoding="utf-8-sig", errors="replace", newline="") as fh:
            reader = csv.reader(fh)
            for row in reader:
                if any(cell.strip() for cell in row):
                    lines.append("\t".join(cell.strip() for cell in row))
        return "\n".join(lines)

    def _read_xlsx(self, path: Path) -> str:
        from openpyxl import load_workbook

        workbook = load_workbook(path, read_only=True, data_only=True)
        sections: list[str] = []

        try:
            for sheet in workbook.worksheets:
                sections.append(f"[Sheet: {sheet.title}]")
                for row in sheet.iter_rows(values_only=True):
                    cells = [str(cell).strip() for cell in row if cell is not None and str(cell).strip()]
                    if cells:
                        sections.append("\t".join(cells))
        finally:
            workbook.close()

        return "\n".join(sections)

    @staticmethod
    def _extract_terms(query: str) -> list[str]:
        terms: list[str] = []

        def add_term(token: str) -> None:
            token = token.strip().lower()
            if len(token) >= 2 and token not in terms:
                terms.append(token)

        for token in re.findall(r"[\u4e00-\u9fff]+|[a-zA-Z0-9]+", query.lower()):
            add_term(token)
            if re.fullmatch(r"[\u4e00-\u9fff]+", token):
                for length in (2, 3, 4):
                    for i in range(len(token) - length + 1):
                        add_term(token[i : i + length])

        return terms
